# 完整可用的 Excel 行事曆 AI 解析器
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import google.generativeai as genai
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import json
import os
from datetime import datetime, timedelta
import re
import pickle

class ExcelCalendarAIParser:
    def __init__(self, gemini_api_key, credentials_file):
        """
        Excel 行事曆解析器
        
        Args:
            gemini_api_key (str): Gemini API 金鑰
            credentials_file (str): Google 憑證檔案路徑 (credentials.json)
        """
        print("🚀 初始化 Excel 行事曆解析器...")
        
        # 設定 Gemini API
        try:
            genai.configure(api_key=gemini_api_key)
            self.model = genai.GenerativeModel('gemini-2.0-flash')
            print("✅ Gemini API 設定成功")
        except Exception as e:
            raise Exception(f"❌ Gemini API 設定失敗: {str(e)}")
        
        # Google Calendar API 設定
        self.SCOPES = ['https://www.googleapis.com/auth/calendar']
        self.credentials_file = credentials_file
        self.calendar_service = None
        
        # 檢查憑證檔案是否存在
        if not os.path.exists(credentials_file):
            raise FileNotFoundError(f"❌ 找不到憑證檔案: {credentials_file}")
        
        print("✅ 初始化完成")

    def setup_google_calendar_api(self):
        """設定 Google Calendar API 認證"""
        print("🔐 開始設定 Google Calendar API...")
        
        creds = None
        token_file = 'token.pickle'
        
        # 檢查是否有已保存的憑證
        if os.path.exists(token_file):
            print("📂 發現已保存的認證token...")
            with open(token_file, 'rb') as token:
                creds = pickle.load(token)
        
        # 如果沒有有效憑證，進行認證流程
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                print("🔄 嘗試刷新過期的認證token...")
                try:
                    creds.refresh(Request())
                    print("✅ Token 刷新成功")
                except Exception as e:
                    print(f"⚠️ Token 刷新失敗: {e}")
                    creds = None
            
            if not creds:
                print("🌐 需要進行 OAuth 認證...")
                print("📝 這將會自動開啟瀏覽器進行 Google 帳號授權")
                
                try:
                    flow = InstalledAppFlow.from_client_secrets_file(
                        self.credentials_file, self.SCOPES)
                    
                    # 使用本地服務器接收認證
                    print("🔗 正在啟動本地認證服務器...")
                    creds = flow.run_local_server(port=0)
                    print("✅ OAuth 認證成功")
                    
                except Exception as e:
                    print(f"❌ 自動認證失敗: {e}")
                    print("💡 嘗試手動認證方法...")
                    
                    try:
                        # 手動認證備用方案
                        flow = InstalledAppFlow.from_client_secrets_file(
                            self.credentials_file, self.SCOPES)
                        
                        # 獲取認證 URL
                        auth_url, _ = flow.authorization_url(prompt='consent')
                        
                        print("\n" + "="*50)
                        print("📋 手動認證步驟：")
                        print("1. 請在瀏覽器中開啟以下 URL：")
                        print(f"   {auth_url}")
                        print("2. 完成 Google 帳號授權")
                        print("3. 複製授權後顯示的認證碼")
                        print("="*50)
                        
                        auth_code = input("請輸入認證碼: ").strip()
                        flow.fetch_token(code=auth_code)
                        creds = flow.credentials
                        print("✅ 手動認證成功")
                        
                    except Exception as e2:
                        raise Exception(f"❌ 所有認證方法都失敗了: {e2}")
            
            # 保存憑證
            with open(token_file, 'wb') as token:
                pickle.dump(creds, token)
            print("💾 認證憑證已保存")
        
        # 建立 Calendar 服務
        self.calendar_service = build('calendar', 'v3', credentials=creds)
        print("✅ Google Calendar API 設定完成")

    def read_excel_file(self, file_path, sheet_name=None):
        """讀取 Excel 檔案，包括合併格處理"""
        print(f"📖 正在讀取 Excel 檔案: {file_path}")
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"❌ 找不到 Excel 檔案: {file_path}")
        
        try:
            # 使用 openpyxl 處理合併格
            workbook = load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"❌ 找不到工作表: {sheet_name}")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            print(f"📄 使用工作表: {worksheet.title}")
            
            # 獲取合併格信息
            merged_ranges = []
            for merged_range in worksheet.merged_cells.ranges:
                merged_ranges.append({
                    'range': str(merged_range),
                    'start_row': merged_range.min_row,
                    'end_row': merged_range.max_row,
                    'start_col': merged_range.min_col,
                    'end_col': merged_range.max_col
                })
            
            # 讀取所有數據
            data = []
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row, col)
                    value = cell.value
                    if value is None:
                        value = ""
                    row_data.append(value)
                data.append(row_data)
            
            result = {
                'data': data,
                'merged_ranges': merged_ranges,
                'max_row': max_row,
                'max_col': max_col,
                'sheet_name': worksheet.title
            }
            
            print(f"✅ Excel 讀取完成: {max_row} 行 × {max_col} 列")
            print(f"🔗 發現 {len(merged_ranges)} 個合併格")
            
            return result
            
        except Exception as e:
            raise Exception(f"❌ 讀取 Excel 檔案失敗: {str(e)}")

    def ai_parse_calendar(self, excel_data):
        """使用 AI 解析行事曆數據"""
        print("🤖 正在使用 AI 解析行事曆數據...")
        
        try:
            # 分析時間列和合併格的對應關係
            time_analysis = self._analyze_time_schedule(excel_data)
            
            # 準備數據字符串（限制長度避免 token 超限）
            data_preview = []
            for i, row in enumerate(excel_data['data'][:25]):  # 取前25行
                if any(cell for cell in row if cell):  # 跳過空行
                    data_preview.append(f"第{i+1}行: {row}")
            
            data_str = "\n".join(data_preview)
            
            # 合併格信息 - 現在包含時間範圍分析
            merged_info = []
            for merged in excel_data['merged_ranges'][:15]:  # 取前15個合併格
                time_range = self._get_time_range_for_merged_cell(merged, time_analysis)
                merged_info.append(f"合併格 {merged['range']}: 第{merged['start_row']}-{merged['end_row']}行，第{merged['start_col']}-{merged['end_col']}列，對應時間：{time_range}")
            
            merged_str = "\n".join(merged_info)
            
            # 建立 AI 提示
            prompt = f"""
你是專業的行事曆數據分析師。請分析以下 Excel 行事曆數據並提取事件信息。

Excel 數據內容：
{data_str}

合併格信息：
{merged_str}

請提取所有有效的行事曆事件，並返回 JSON 格式的數據。

每個事件必須包含：
- title: 事件名稱（必需）
- start_date: 開始日期，格式 YYYY-MM-DD（必需）
- start_time: 開始時間，格式 HH:MM，預設 "09:00"
- end_date: 結束日期，格式 YYYY-MM-DD（必需）
- end_time: 結束時間，格式 HH:MM，預設 "18:00"
- description: 事件描述（可選）
- location: 地點（可選）

重要提示：
1. **合併格時間對應**：合併格信息中已經包含了該合併格對應的完整時間範圍
   - 當看到合併格資訊顯示「對應時間：08:25~10:05」時，請直接使用這個時間範圍
   - 不要只取合併格最上面一格的時間，要使用完整的時間範圍
2. 仔細分析日期和時間格式，可能有各種表示方式
3. 忽略空白或無意義的數據
4. 如果日期不完整，請根據上下文推測完整日期
5. **課程表的合併格表示連堂課程**：
   - 如果一個課程名稱出現在合併格中，它就是一個跨多個時段的連堂課程
   - 請使用合併格信息中提供的完整時間範圍
   - 例如：合併格顯示「數學，對應時間：08:25~10:05」，則事件時間就是08:25~10:05
6. 對於非合併格的單獨課程，使用該格對應的單個時段時間
7. **重要：請只返回純 JSON 陣列，不要包含任何其他文字、解釋、代碼塊或markdown格式**
8. **禁止**返回Python代碼或任何程式碼，只要純JSON格式

輸出格式要求：
- 必須是有效的JSON陣列格式
- 不要使用```json或```包裝
- 不要有任何前後綴文字說明
- 直接輸出JSON陣列

正確的輸出範例：
[
  {{
    "title": "重要會議",
    "start_date": "2024-12-01",
    "start_time": "10:00", 
    "end_date": "2024-12-01",
    "end_time": "12:00",
    "description": "討論重要事項",
    "location": "會議室A"
  }},
  {{
    "title": "數學課（連堂）",
    "start_date": "2024-12-02",
    "start_time": "08:25", 
    "end_date": "2024-12-02",
    "end_time": "10:05",
    "description": "橫跨第一、二節課的連堂課程",
    "location": "教室101"
  }}
]
"""

            # 呼叫 AI
            response = self.model.generate_content(prompt)
            response_text = response.text.strip()
            
            print(f"🤖 AI 原始回應長度: {len(response_text)} 字符")
            print(f"🤖 AI 回應前100字符: {response_text[:100]}...")
            
            # 更強化的回應清理邏輯
            # 移除可能的代碼塊標記
            if response_text.startswith('```'):
                # 找到第一個換行後的內容
                lines = response_text.split('\n')
                if len(lines) > 1:
                    response_text = '\n'.join(lines[1:])
                else:
                    response_text = response_text[3:]  # 移除```
            
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            
            # 移除其他可能的標記
            prefixes_to_remove = ['```json', '```python', 'json', 'python']
            for prefix in prefixes_to_remove:
                if response_text.startswith(prefix):
                    response_text = response_text[len(prefix):]
                    break
            
            response_text = response_text.strip()
            
            # 檢查是否為JSON格式
            if not response_text.startswith('['):
                print("⚠️ AI回應可能不是有效的JSON格式，嘗試提取JSON部分...")
                
                # 嘗試找到JSON陣列部分
                start_idx = response_text.find('[')
                end_idx = response_text.rfind(']')
                
                if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                    response_text = response_text[start_idx:end_idx+1]
                    print(f"✅ 提取到JSON部分: {len(response_text)} 字符")
                else:
                    print("❌ 無法找到有效的JSON陣列格式")
                    return []
            
            print(f"🔍 處理後的回應前100字符: {response_text[:100]}...")
            
            # 解析 JSON
            events = json.loads(response_text)
            
            print(f"✅ AI 解析完成，找到 {len(events)} 個事件")
            
            # 顯示事件預覽
            if events:
                print("📅 事件預覽：")
                for i, event in enumerate(events[:5], 1):
                    print(f"  {i}. {event['title']} - {event['start_date']} {event['start_time']}")
                if len(events) > 5:
                    print(f"  ... 還有 {len(events)-5} 個事件")
            
            # 後處理：合併相同課程的連續時段
            events = self._merge_consecutive_events(events)
            
            return events
            
        except json.JSONDecodeError as e:
            print(f"❌ AI 回應的 JSON 格式錯誤: {e}")
            print(f"AI 原始回應: {response.text[:300]}...")
            return []
        except Exception as e:
            print(f"❌ AI 解析失敗: {e}")
            return []

    def _merge_consecutive_events(self, events):
        """合併相同課程的連續時段"""
        if not events:
            return events
        
        print("🔄 檢查並合併連續時段...")
        
        # 按日期和開始時間排序
        events_sorted = sorted(events, key=lambda x: (x['start_date'], x['start_time']))
        
        merged_events = []
        i = 0
        
        while i < len(events_sorted):
            current_event = events_sorted[i].copy()
            
            # 尋找相同課程名稱的連續事件
            j = i + 1
            while j < len(events_sorted):
                next_event = events_sorted[j]
                
                # 檢查是否為相同課程且在同一天
                if (current_event['title'].strip() == next_event['title'].strip() and 
                    current_event['start_date'] == next_event['start_date']):
                    
                    # 檢查時段是否連續（間隔少於30分鐘）
                    current_end_time = self._time_to_minutes(current_event['end_time'])
                    next_start_time = self._time_to_minutes(next_event['start_time'])
                    
                    if next_start_time - current_end_time <= 30:  # 30分鐘內視為連續
                        # 合併時段
                        current_event['end_time'] = next_event['end_time']
                        current_event['end_date'] = next_event['end_date']
                        
                        # 更新描述
                        if not current_event.get('description'):
                            current_event['description'] = ""
                        if "連堂" not in current_event['description']:
                            current_event['description'] += " (連堂課程)" if current_event['description'] else "連堂課程"
                        
                        print(f"  🔗 合併課程: {current_event['title']} {current_event['start_time']}-{current_event['end_time']}")
                        j += 1
                    else:
                        break
                else:
                    break
            
            merged_events.append(current_event)
            i = j if j > i + 1 else i + 1
        
        print(f"✅ 時段合併完成，從 {len(events)} 個事件合併為 {len(merged_events)} 個事件")
        return merged_events
    
    def _time_to_minutes(self, time_str):
        """將時間字符串轉換為分鐘數（從午夜開始計算）"""
        try:
            hours, minutes = map(int, time_str.split(':'))
            return hours * 60 + minutes
        except:
            return 0

    def _analyze_time_schedule(self, excel_data):
        """分析課程表的時間結構"""
        print("🕐 分析時間表結構...")
        
        time_schedule = {}
        data = excel_data['data']
        
        # 尋找時間列（通常在左邊幾列）
        for col_idx in range(min(3, excel_data['max_col'])):  # 檢查前3列
            for row_idx in range(excel_data['max_row']):
                cell_value = str(data[row_idx][col_idx]).strip()
                
                # 檢查是否包含時間格式
                if self._is_time_format(cell_value):
                    time_schedule[row_idx + 1] = {  # Excel行號從1開始
                        'time_text': cell_value,
                        'parsed_time': self._parse_time_range(cell_value),
                        'column': col_idx + 1
                    }
        
        print(f"✅ 找到 {len(time_schedule)} 個時間段")
        return time_schedule
    
    def _is_time_format(self, text):
        """檢查文字是否包含時間格式"""
        import re
        
        # 檢查各種時間格式
        time_patterns = [
            r'\d{1,2}:\d{2}',  # 08:30
            r'\d{1,2}：\d{2}',  # 08：30 (中文冒號)
            r'第\d+節',  # 第1節
            r'\d+節',  # 1節
            r'上午|下午|早上|中午|晚上',  # 時間描述
        ]
        
        for pattern in time_patterns:
            if re.search(pattern, text):
                return True
        return False
    
    def _parse_time_range(self, time_text):
        """解析時間範圍文字"""
        import re
        
        # 提取時間
        time_matches = re.findall(r'(\d{1,2}):(\d{2})', time_text)
        if len(time_matches) >= 2:
            # 有開始和結束時間
            start_hour, start_min = int(time_matches[0][0]), int(time_matches[0][1])
            end_hour, end_min = int(time_matches[1][0]), int(time_matches[1][1])
            return {
                'start_time': f"{start_hour:02d}:{start_min:02d}",
                'end_time': f"{end_hour:02d}:{end_min:02d}"
            }
        elif len(time_matches) == 1:
            # 只有一個時間，假設為開始時間
            start_hour, start_min = int(time_matches[0][0]), int(time_matches[0][1])
            end_hour = start_hour + 1  # 假設一小時的課程
            return {
                'start_time': f"{start_hour:02d}:{start_min:02d}",
                'end_time': f"{end_hour:02d}:{start_min:02d}"
            }
        
        # 解析節次
        period_match = re.search(r'第?(\d+)節', time_text)
        if period_match:
            period = int(period_match.group(1))
            # 假設每節課50分鐘，從8:00開始
            start_hour = 8 + (period - 1)
            end_hour = start_hour + 1
            return {
                'start_time': f"{start_hour:02d}:00",
                'end_time': f"{end_hour:02d}:00"
            }
        
        return None
    
    def _get_time_range_for_merged_cell(self, merged_cell, time_schedule):
        """獲取合併儲存格對應的時間範圍"""
        start_row = merged_cell['start_row']
        end_row = merged_cell['end_row']
        
        # 收集這個範圍內所有的時間
        time_ranges = []
        for row in range(start_row, end_row + 1):
            if row in time_schedule:
                parsed_time = time_schedule[row]['parsed_time']
                if parsed_time:
                    time_ranges.append(parsed_time)
        
        if not time_ranges:
            return "無法識別時間"
        
        # 找到最早的開始時間和最晚的結束時間
        earliest_start = min(tr['start_time'] for tr in time_ranges)
        latest_end = max(tr['end_time'] for tr in time_ranges)
        
        return f"{earliest_start}~{latest_end}"

    def create_calendar_events(self, events):
        """在 Google Calendar 中建立事件"""
        if not self.calendar_service:
            raise Exception("❌ Google Calendar 服務未設定")
        
        print(f"📅 開始同步 {len(events)} 個事件到 Google Calendar...")
        
        success_count = 0
        failed_count = 0
        failed_events = []
        
        for i, event_data in enumerate(events, 1):
            try:
                print(f"正在建立事件 {i}/{len(events)}: {event_data['title']}")
                
                # 建立 Google Calendar 事件
                calendar_event = {
                    'summary': event_data['title'],
                    'start': {
                        'dateTime': f"{event_data['start_date']}T{event_data['start_time']}:00",
                        'timeZone': 'Asia/Taipei',
                    },
                    'end': {
                        'dateTime': f"{event_data['end_date']}T{event_data['end_time']}:00",
                        'timeZone': 'Asia/Taipei',
                    },
                }
                
                # 添加可選欄位
                if event_data.get('description'):
                    calendar_event['description'] = event_data['description']
                if event_data.get('location'):
                    calendar_event['location'] = event_data['location']
                
                # 建立事件
                result = self.calendar_service.events().insert(
                    calendarId='primary',
                    body=calendar_event
                ).execute()
                
                success_count += 1
                print(f"  ✅ 成功建立事件")
                
            except Exception as e:
                failed_count += 1
                failed_events.append({
                    'event': event_data,
                    'error': str(e)
                })
                print(f"  ❌ 建立失敗: {str(e)}")
        
        return {
            'success': success_count,
            'failed': failed_count,
            'total': len(events),
            'failed_events': failed_events
        }

    def process_calendar(self, excel_file, sheet_name=None):
        """完整處理流程"""
        print("=" * 60)
        print("🎯 Excel 行事曆 AI 同步工具")
        print("=" * 60)
        
        try:
            # 1. 讀取 Excel
            print("\n📖 步驟 1: 讀取 Excel 檔案")
            excel_data = self.read_excel_file(excel_file, sheet_name)
            
            # 2. AI 解析
            print("\n🤖 步驟 2: AI 智能解析")
            events = self.ai_parse_calendar(excel_data)
            
            if not events:
                print("❌ 沒有找到任何有效事件")
                return {
                    'status': 'no_events',
                    'message': '沒有找到任何有效的行事曆事件'
                }
            
            # 3. 同步到 Google Calendar
            print("\n📅 步驟 3: 同步到 Google Calendar")
            sync_result = self.create_calendar_events(events)
            
            # 顯示結果
            print("\n" + "=" * 60)
            print("🎉 處理完成！")
            print(f"📊 統計結果：")
            print(f"   總事件數: {sync_result['total']}")
            print(f"   成功同步: {sync_result['success']}")
            print(f"   同步失敗: {sync_result['failed']}")
            
            if sync_result['failed'] > 0:
                print("\n❌ 失敗的事件：")
                for failed in sync_result['failed_events'][:3]:
                    print(f"   - {failed['event']['title']}: {failed['error']}")
            
            print("=" * 60)
            
            return {
                'status': 'success',
                'events': events,
                'sync_result': sync_result
            }
            
        except Exception as e:
            print(f"\n❌ 處理失敗: {str(e)}")
            return {
                'status': 'error',
                'message': str(e)
            }


# 簡單使用範例
def main():
    """主程式 - 一鍵執行"""
    
    # ⚠️ 請替換以下參數 ⚠️
    GEMINI_API_KEY = "your_gemini_api_key_here"    # 您的 Gemini API 金鑰
    CREDENTIALS_FILE = "credentials.json"          # Google 憑證檔案路徑
    EXCEL_FILE = "calendar.xlsx"                   # Excel 行事曆檔案路徑
    
    print("🚀 啟動 Excel 行事曆 AI 同步工具")
    
    try:
        # 建立解析器
        parser = ExcelCalendarAIParser(
            gemini_api_key=GEMINI_API_KEY,
            credentials_file=CREDENTIALS_FILE
        )
        
        # 設定 Google Calendar API
        parser.setup_google_calendar_api()
        
        # 處理行事曆
        result = parser.process_calendar(EXCEL_FILE)
        
        if result['status'] == 'success':
            print("\n🎊 太棒了！您的 Excel 行事曆已成功同步到 Google Calendar！")
            print("📱 現在可以在手機和電腦上的 Google Calendar 中查看您的事件了")
        else:
            print(f"\n😔 處理過程遇到問題: {result.get('message', '未知錯誤')}")
    
    except Exception as e:
        print(f"\n💥 程式執行錯誤: {str(e)}")
        print("\n🔧 請檢查：")
        print("1. Gemini API 金鑰是否正確")
        print("2. credentials.json 檔案是否存在")
        print("3. Excel 檔案路徑是否正確")
        print("4. 網路連線是否正常")


if __name__ == "__main__":
    main()